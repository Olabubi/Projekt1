import json
from datetime import datetime
from openpyxl import Workbook, load_workbook
from fpdf import FPDF


class Task:
    def __init__(self, description, due_date=None, priority='Low', completed=False, completion_date=''):
        self.description = description
        self.due_date = due_date
        self.priority = priority
        self.completed = completed
        self.completion_date = completion_date

    def mark_completed(self):
        self.completed = True
        self.completion_date = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

    def mark_pending(self):
        self.completed = False
        self.completion_date = ''

    def __repr__(self):
        return f"{'[x]' if self.completed else '[ ]'} {self.description} - Due: {self.due_date} - Priority: {self.priority} - Completed: {self.completion_date}"


class TaskManager:
    def __init__(self, file_path='tasks.json', exported_tasks_file='exported_tasks.json'):
        self.tasks = []
        self.file_path = file_path
        self.exported_tasks_file = exported_tasks_file
        self.exported_tasks = self.load_exported_tasks()
        self.load_tasks()
        self.wb = Workbook()
        self.ws = self.wb.active

    def add_task(self, description, due_date, priority):
        task = Task(description, due_date, priority)
        self.tasks.append(task)
        self.save_tasks()

    def delete_task(self, index):
        if 0 <= index < len(self.tasks):
            del self.tasks[index]
            self.save_tasks()

    def mark_task_completed(self, index):
        if 0 <= index < len(self.tasks):
            self.tasks[index].mark_completed()
            self.save_tasks()

    def unmark_task_completed(self, index):
        if 0 <= index < len(self.tasks):
            self.tasks[index].mark_pending()
            self.save_tasks()

    def export_to_excel(self, tasks):
        headers = ["description", "due_date", "priority", "completed", "completion_date"]

        # Load or create a workbook
        try:
            self.wb = load_workbook("tasks.xlsx")
            self.ws = self.wb.active
        except FileNotFoundError:
            self.ws.append(headers)

        # Add new tasks
        with open(self.file_path, 'r') as file:
            data = json.load(file)
            for task in data:
                if task not in self.exported_tasks:
                    row = [task.get(header, "") for header in headers]
                    self.ws.append(row)
                    self.exported_tasks.append(task)

        self.wb.save("tasks.xlsx")
        self.save_exported_tasks()

    def export_to_pdf(self, tasks):
        pdf = FPDF()
        pdf.add_page()
        pdf.set_font("Arial", size=12)

        headers = ["Description", "Due Date", "Priority", "Completed Date"]
        col_widths = [60, 30, 30, 40]

        # nagłówki do pdf
        for header, width in zip(headers, col_widths):
            pdf.cell(width, 10, header, border=1, ln=False, align="C")

        pdf.ln()

        # dodaje zadania do pdf
        for task in tasks:
            pdf.cell(col_widths[0], 10, task.description, border=1)
            pdf.cell(col_widths[1], 10, task.due_date, border=1)
            pdf.cell(col_widths[2], 10, task.priority, border=1)
            pdf.cell(col_widths[3], 10, task.completion_date if task.completion_date else "", border=1)
            pdf.ln()

        pdf_file_path = f"tasks_{datetime.now().strftime('%Y-%m-%d %H_%M_%S')}.pdf"
        pdf.output(pdf_file_path)

    def view_tasks(self, completed=None):
        if completed is None:
            return self.tasks
        return [task for task in self.tasks if task.completed == completed]

    def save_tasks(self):
        with open(self.file_path, 'w') as file:
            json.dump([task.__dict__ for task in self.tasks], file)

    def load_tasks(self):
        try:
            with open(self.file_path, 'r') as file:
                tasks_data = json.load(file)
                self.tasks = [Task(**data) for data in tasks_data]
        except FileNotFoundError:
            self.tasks = []

    def load_exported_tasks(self):
        try:
            with open(self.exported_tasks_file, 'r') as file:
                return json.load(file)
        except FileNotFoundError:
            return []

    def save_exported_tasks(self):
        with open(self.exported_tasks_file, 'w') as file:
            json.dump(self.exported_tasks, file)

    def cancel_task(self, index):
        if 0 <= index < len(self.tasks):
            del self.tasks[index]
            self.save_tasks()

    def change_due_date(self, index, new_due_date):
        if index is not None and new_due_date is not None:
            if isinstance(index, int) and 0 <= index < len(self.tasks):
                print("Before:", self.tasks[index].due_date)  # Debugging print
                self.tasks[index].due_date = new_due_date
                print("After:", self.tasks[index].due_date)  # Debugging print
                self.save_tasks()
